Задание: автоматизация ABC анализа продаж при помощи python
    
import openpyxl

wb_obj = openpyxl.load_workbook("pyton.xlsx")
sheet_obj = wb_obj.active

i=3
n=(sheet_obj.max_row)
print(n)
virfer = 0

while i < n:
    t= sheet_obj.cell(row=i, column = 3)
    print(t.value)
    if t.value!=None:
        virfer=virfer+t.value
        i=i + 1
        
    else:
        i = i + 1
print('выручка за ферваль' + ' составила ' + str(virfer))

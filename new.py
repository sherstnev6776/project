Задание: автоматизация ABC анализа продаж при помощи python
 import openpyxl

wb_obj = openpyxl.load_workbook("pyton.xlsx")
sheet_obj = wb_obj.active
u=3
i=3
o=3
p=3
l=3
j=3
s = 3
v = 3
n=(sheet_obj.max_row)
stroki=(sheet_obj.max_column)
print(n)
virfer = 0
virmart = 0
virapr = 0
virmay = 0
virtov = 0
virtov1 = 0 
print(stroki)



maxi=0
#   расчет общих продаж
while j < stroki and u < n:
    vit1 = sheet_obj.cell(row = u, column = j)
    if vit1.value!= None:
        virtov1 = virtov1 + vit1.value
        j = j + 2
    else:
        j = j + 2
    if j == stroki or j > stroki:
        u = u+1
        j = 3
#Расчет суммы продаж по каждому товару за весь период        
dolya = 0       
while s < stroki and v < n:
    vit = sheet_obj.cell(row = v, column = s)
    tovar = sheet_obj.cell(row = v, column = 1)
    if vit.value!= None:
        virtov = virtov + vit.value
        s = s + 2
    else:
        s = s + 2
    if s == stroki or s > stroki:
        v = v+1
        s = 3
        dolya = virtov / virtov1 * 100
        print(str(tovar.value) + 'продано на ' + str(virtov) + ' рублей, что составляет' + "{0:.2f}".format(dolya) + ' % от общих продаж ')
        print()
        virtov = 0 




print(virtov1) # выручка общая
        

'''
while i < n:
    t= sheet_obj.cell(row=i, column = 3)
    #print(t.value)
    if t.value!=None:
        virfer=virfer+t.value
        i= i + 1
    else:
        i = i + 1

while o < n:
    m= sheet_obj.cell(row=o, column = 5)
    #print(m.value)
    if m.value!=None:
        virmart=virmart+m.value
        o= o + 1
    else:
        o= o + 1
       
while p < n:
    a= sheet_obj.cell(row=p, column = 7)
    #print(a.value)
    if a.value!=None:
        virapr=virapr+a.value
        p= p + 1
    else:
        p = p + 1

while l < n:
    ma= sheet_obj.cell(row=l, column = 9)
    #print(ma.value)
    if ma.value!=None:
        virmay=virmay+t.value
        l= l + 1
    else:
        l = l + 1
'''
virvse = virfer + virmart + virapr + virmay
      
print('выручка за ферваль' + ' составила ' + str(virfer))
print('выручка за март' + ' составила ' + str(virmart))
print('выручка за апрель' + ' составила ' + str(virapr))
print('выручка за май' + ' составила ' + str(virmay))
print('всего' + 'выручка' + ' составила ' + str(virvse))
